from __future__ import annotations

import argparse
import csv
import difflib
import json
import logging
import re
import sys
import unicodedata
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable, List, Sequence, Set
from uuid import uuid4

from dateutil import parser as date_parser
import pandas as pd
from openpyxl import Workbook, load_workbook

from preprocessamento.documents import gather_texts, document_priority
from preprocessamento.inputs import PreparedInput, resolve_input_paths
from .doc_classifier import DocumentBucket, classify_document

_PERITO_CATALOG_PATH = Path("outputs/banco-peritos/peritos_catalogo_final.csv")
_PERITO_NAME_SET: set[str] | None = None

_PERITO_CATALOG_PATH = Path("outputs/banco-peritos/peritos_catalogo_final.csv")
_PERITO_NAME_SET: set[str] | None = None

BUCKET_ORDER = [
    DocumentBucket.PRINCIPAL,
    DocumentBucket.APOIO,
    DocumentBucket.LAUDO,
    DocumentBucket.OUTRO,
]

BUCKET_BASE_REQUIREMENTS: dict[DocumentBucket, set[str]] = {
    DocumentBucket.PRINCIPAL: {
        "PROCESSO Nº",
        "PROCESSO ADMIN. Nº",
        "JUÍZO",
        "COMARCA",
        "VALOR ARBITRADO",
        "PROMOVENTE",
        "PROMOVIDO",
        "PERITO",
        "CPF/CNPJ",
    },
    DocumentBucket.APOIO: {
        "DATA DA REQUISIÇÃO",
        "DATA ADIANTAMENTO",
        "Data da Autorização da Despesa",
        "ESPÉCIE DE PERÍCIA",
        "Fator",
        "Valor Tabelado Anexo I - Tabela I",
    },
    DocumentBucket.LAUDO: set(),
    DocumentBucket.OUTRO: set(),
}


def _build_bucket_requirements() -> dict[DocumentBucket, set[str]]:
    cumulative: dict[DocumentBucket, set[str]] = {}
    collected: set[str] = set()
    for bucket in BUCKET_ORDER:
        collected = collected | BUCKET_BASE_REQUIREMENTS.get(bucket, set())
        cumulative[bucket] = set(collected)
    return cumulative


BUCKET_REQUIREMENTS = _build_bucket_requirements()

INVALID_EXCEL_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
LOG_DIR = Path(__file__).resolve().parents[2] / "logs" / "extract"
LOGGER = logging.getLogger("extract_reports")


def _generate_run_id() -> str:
    return f"extract-{datetime.now():%Y%m%d-%H%M%S}-{uuid4().hex[:6]}"


def _state_path(run_id: str) -> Path:
    return LOG_DIR / f"{run_id}.state.json"


def _load_state(run_id: str) -> dict:
    path = _state_path(run_id)
    if not path.exists():
        raise SystemExit(f"Checkpoint não encontrado para run-id {run_id} em {path}")
    return json.loads(path.read_text(encoding="utf-8"))


def _save_state(run_id: str, state: dict) -> None:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    path = _state_path(run_id)
    path.write_text(json.dumps(state, indent=2, ensure_ascii=False), encoding="utf-8")


def _setup_logger(run_id: str) -> Path:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_path = LOG_DIR / f"{run_id}.log"
    handler = logging.FileHandler(log_path, encoding="utf-8")
    handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    LOGGER.handlers.clear()
    LOGGER.addHandler(handler)
    LOGGER.setLevel(logging.INFO)
    LOGGER.propagate = False
    return log_path


def _cleanup_old_logs(days: int) -> None:
    if days <= 0 or not LOG_DIR.exists():
        return
    cutoff = datetime.now() - timedelta(days=days)
    for pattern in ("extract-*.log", "extract-*.state.json"):
        for file in LOG_DIR.glob(pattern):
            try:
                if datetime.fromtimestamp(file.stat().st_mtime) < cutoff:
                    file.unlink()
            except Exception:
                continue


def _log(message: str) -> None:
    print(message)
    if LOGGER.handlers:
        LOGGER.info(message)


def _print_header(run_id: str, log_path: Path, total: int) -> None:
    banner = "=" * 72
    print(banner)
    print(f"Execução ID: {run_id}")
    print(f"Log:        {log_path}")
    print(f"Total de arquivos: {total}")
    print(banner)


def _log_progress(completed: int, total: int, name: str | None = None) -> None:
    if total <= 0:
        return
    pct = completed / total
    msg = f"{completed}/{total} ({pct*100:5.1f}%)"
    if name:
        msg = f"{msg} - {name}"
    _log(msg)


def _finish_progress() -> None:
    print()  # move to next line


PROMOVENTE_LABELS = ("promovente", "parte autora", "autor", "requerente")
PROMOVIDO_LABELS = (
    "promovido",
    "requerido",
    "requerida",
    "parte ré",
    "réu",
    "ré",
    "demandado",
    "demandada",
    "parte contrária",
    "executado",
    "executada",
)
PERITO_LABELS = ("perito", "perita")
CPF_LABELS = ("cpf", "cnpj")
ESPECIALIDADE_LABELS = ("especialidade", "área de atuação")
ESPECIE_LABELS = ("espécie de perícia", "espécie", "tipo de perícia")
NATUREZA_LABELS = ("natureza dos honorários", "natureza dos honorarios", "natureza do serviço", "natureza da perícia")
FATOR_LABELS = ("fator",)
VALOR_TABELA_LABELS = ("valor tabelado", "valor tabela")
VALOR_ARBITRADO_LABELS = ("valor arbitrado", "valor da perícia", "honorários")
CHECAGEM_LABELS = ("checagem",)
DATA_ADIANTAMENTO_LABELS = ("data adiantamento", "data do adiantamento")
CHECAGEM_ADIANT_LABELS = ("checagem adiantamento",)
DATA_AUTORIZACAO_LABELS = ("data da autorização", "autorização da despesa")
SALDO_LABELS = ("saldo a receber",)

PROCESSO_NUM_PATTERN = re.compile(r"\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}")
PROCESSO_ADMIN_PATTERN = re.compile(r"20\d{8}")
PROCESSO_ADMIN_LEGACY_PATTERN = re.compile(r"20\d{2}[\.\-/ ]?\d{3,6}")
PROCESSO_SEI_PATTERN = re.compile(r"\d{5,7}-\d{2}\.\d{4}\.\d\.\d{2}")
PROCESSO_SEI_LOOSE_PATTERN = re.compile(
    r"(\d{5,7})\s*-\s*(\d{2})[\.\-/ ]*(\d{4})[\.\-/ ]*(\d)[\.\-/ ]*(\d{2})(?![\.\-/ ]?\d{3,4})"
)
PROCESSO_SEI_LOOSE_PATTERN = re.compile(
    r"(\d{5,7})\s*-\s*(\d{2})[.\-\s]*(\d{4})[.\-\s]*(\d)[.\-\s]*(\d{2})(?![.\-]\d{3,4})"
)
CPF_PATTERN = re.compile(r"\d{3}\.\d{3}\.\d{3}-\d{2}")
CNPJ_PATTERN = re.compile(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}")
PERITO_PARAGRAPH_PATTERN = re.compile(
    r"Perit[oa]\s+(?P<esp>[^,]+),\s*(?P<nome>[A-Za-zÀ-ÿ' ]+),\s*CPF\s*(?P<cpf>\d{3}\.\d{3}\.\d{3}-\d{2})",
    re.IGNORECASE,
)
DATE_PATTERN = re.compile(
    r"\b\d{1,2}[-/ ]?(?:jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)\.?-?\.?\d{2,4}\b",
    re.IGNORECASE,
)
DATE_NUMERIC_PATTERN = re.compile(r"\b\d{1,2}[./-]\d{1,2}[./-]\d{2,4}\b")
JUÍZO_PATTERN = re.compile(r"\b\d+ª\s+Vara[^\n]+", re.IGNORECASE)
COMARCA_PATTERN = re.compile(r"Comarca\s+de\s+[A-Za-zÀ-ÿ ]+", re.IGNORECASE)
PARTES_REGEX = re.compile(
    r"(?:movid[oa]|propost[oa]|promov[ei]d[oa])\s+por\s+(?P<promovente>[^,\n]+?)"
    r"(?:,|\s+CPF|\s+CNPJ|\se[mn]\s+face)"
    r".*?em\s+face\s+(?:de|da|do|dos|das)\s+(?P<promovido>[^,\n]+)",
    re.IGNORECASE | re.DOTALL,
)
ESPECIE_INLINE_PATTERN = re.compile(
    r"(?:esp[eé]cie|tipo)\s+de\s+per[ií]cia\s*[:\-]?\s*([^\n;]+)", re.IGNORECASE
)
NATUREZA_HONORARIOS_PATTERN = re.compile(
    r"natureza\s+dos\s+honor[aá]rios\s*[:\-]?\s*([^\n;]+)", re.IGNORECASE
)
NATUREZA_SERVICO_PATTERN = re.compile(
    r"natureza\s+do\s+servi[cç]o\s*[:\-]?\s*([^\n;]+)", re.IGNORECASE
)
PROMOVIDO_EM_FACE_PATTERN = re.compile(
    r"em\s+(?:face|desfavor)\s+(?:do|da|de|dos|das|d[oa]s?)\s+([^\n;,]+)",
    re.IGNORECASE,
)
PROMOVIDO_CONTRA_PATTERN = re.compile(
    r"contra\s+(?:o|a|os|as)\s+([^\n;,]+)",
    re.IGNORECASE,
)
PROMOVIDO_LABEL_PATTERN = re.compile(
    r"(?:réu|ré|requerid[ao]s?|demandad[ao]s?|parte\s+contr[áa]ria)\s*[:\-]\s*([^\n;,]+)",
    re.IGNORECASE,
)

REQUISITION_KEYWORDS = ("requisição", "requerimento", "solicitação")
CITY_KEYWORDS = (
    "campina grande",
    "joão pessoa",
    "patos",
    "sousa",
    "cabedelo",
    "guarabira",
    "areia",
)

CNJ_CONTEXT_KEYWORDS = (
    "nos autos",
    "processo",
    "perito",
    "honor",
    "autor",
    "promovente",
    "promovido",
    "requisição",
)

COLUMNS = [
    "Nº DE PERÍCIAS",
    "DATA DA REQUISIÇÃO",
    "PROCESSO ADMIN. Nº",
    "JUÍZO",
    "COMARCA",
    "PROCESSO Nº",
    "PROMOVENTE",
    "PROMOVIDO",
    "PERITO",
    "CPF/CNPJ",
    "ESPECIALIDADE",
    "ESPÉCIE DE PERÍCIA",
    "Fator",
    "Valor Tabelado Anexo I - Tabela I",
    "VALOR ARBITRADO - DE",
    "VALOR ARBITRADO - CM",
    "VALOR ARBITRADO",
    "CHECAGEM",
    "DATA ADIANTAMENTO",
    "R$",
    "%",
    "CHECAGEM ADIANTAMENTO",
    "Data da Autorização da Despesa",
    "SALDO A RECEBER",
    "ARQUIVO_ORIGEM",
    "OBSERVACOES",
]

HONORARIOS_TABLE: list[dict[str, str]] = []
HONORARIOS_INDEX: dict[str, dict[str, str]] = {}
HONORARIOS_BY_ID: dict[str, dict[str, str]] = {}
HONORARIOS_ALIAS: list[dict[str, object]] = []


def _normalize_key(value: str) -> str:
    if not value:
        return ""
    normalized = unicodedata.normalize("NFD", value)
    normalized = normalized.encode("ascii", "ignore").decode("ascii")
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized.lower()).strip()
    return normalized


def _safe_excel_value(value):
    if isinstance(value, str):
        return INVALID_EXCEL_CHARS.sub("", value)
    return value


def _load_honorarios_table() -> None:
    base_dir = Path(__file__).resolve().parents[2]
    path = base_dir / "docs" / "tabela_honorarios.csv"
    if not path.exists():
        return
    try:
        with path.open("r", encoding="utf-8") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                desc = row.get("DESCRICAO", "").strip()
                if not desc:
                    continue
                row["VALOR"] = row.get("VALOR", "").strip()
                HONORARIOS_TABLE.append(row)
                HONORARIOS_INDEX[_normalize_key(desc)] = row
                row_id = row.get("ID", "").strip()
                if row_id:
                    HONORARIOS_BY_ID[row_id] = row
    except Exception:
        HONORARIOS_TABLE.clear()
        HONORARIOS_INDEX.clear()
        HONORARIOS_BY_ID.clear()


_load_honorarios_table()


def _load_honorarios_aliases() -> None:
    base_dir = Path(__file__).resolve().parents[2]
    path = base_dir / "docs" / "honorarios_aliases.json"
    if not path.exists():
        return
    try:
        data = json.loads(path.read_text("utf-8"))
    except Exception:
        return
    for item in data:
        target_id = str(item.get("target_id", "")).strip()
        entry = HONORARIOS_BY_ID.get(target_id)
        if not entry:
            continue
        keywords = [kw for kw in (_normalize_key(k) for k in item.get("keywords", [])) if kw]
        if not keywords:
            continue
        HONORARIOS_ALIAS.append({"keywords": keywords, "entry": entry})


_load_honorarios_aliases()


@dataclass
class ExtractionResult:
    data: dict[str, str] = field(default_factory=dict)
    observations: List[str] = field(default_factory=list)
    sources: dict[str, str] = field(default_factory=dict)
    meta: dict[str, dict] = field(default_factory=dict)
    candidates: dict[str, list[dict]] = field(default_factory=dict)

    def update_from(self, other: "ExtractionResult", source_name: str) -> None:
        for key, value in other.data.items():
            if not value:
                continue
            current_meta = self.meta.get(key)
            incoming_meta = other.meta.get(key)
            current_weight = current_meta.get("weight", 0) if current_meta else 0
            incoming_weight = incoming_meta.get("weight", 0) if incoming_meta else 0
            if not self.data.get(key) or incoming_weight > current_weight:
                self.data[key] = value
                self.sources[key] = source_name
                if incoming_meta:
                    self.meta[key] = incoming_meta
        for field, entries in other.candidates.items():
            self.candidates.setdefault(field, []).extend(entries)
        for obs in other.observations:
            if obs not in self.observations:
                self.observations.append(obs)

    def to_row(self, index: int, zip_name: str) -> list[str]:
        row = []
        for column in COLUMNS:
            if column == "Nº DE PERÍCIAS":
                row.append(f"{index:02d}")
            elif column == "ARQUIVO_ORIGEM":
                row.append(zip_name)
            elif column == "OBSERVACOES":
                row.append("; ".join(self.observations))
            else:
                row.append(self.data.get(column, ""))
        return row


@dataclass
class DocumentText:
    name: str
    text: str
    sei_numbers: Set[str] = field(default_factory=set)
    judicial_numbers: Set[str] = field(default_factory=set)
    admin_numbers: Set[str] = field(default_factory=set)
    cnj_counts: dict[str, int] = field(default_factory=dict)
    cnj_context: dict[str, int] = field(default_factory=dict)
    cnj_display: dict[str, str] = field(default_factory=dict)
    importance: int = 1
    bucket: DocumentBucket = DocumentBucket.OUTRO

    @property
    def has_ids(self) -> bool:
        return bool(self.sei_numbers or self.judicial_numbers or self.admin_numbers)


@dataclass
class ProcessContext:
    expected_sei: Set[str] = field(default_factory=set)
    expected_sei_display: dict[str, str] = field(default_factory=dict)
    accepted_docs: List[DocumentText] = field(default_factory=list)
    skipped_docs: List[str] = field(default_factory=list)
    sei_numbers: Set[str] = field(default_factory=set)
    judicial_numbers: Set[str] = field(default_factory=set)
    admin_numbers: Set[str] = field(default_factory=set)

    def has_anchor(self) -> bool:
        return bool(self.sei_numbers or self.judicial_numbers or self.admin_numbers or self.expected_sei)

    def register(self, doc: DocumentText) -> None:
        self.accepted_docs.append(doc)
        self.sei_numbers.update(doc.sei_numbers)
        self.judicial_numbers.update(doc.judicial_numbers)
        self.admin_numbers.update(doc.admin_numbers)




def _normalize_digits(value: str) -> str:
    return re.sub(r"\D", "", value or "")


def _normalize_sei_number(value: str) -> str:
    return _normalize_digits(value)


def _normalize_judicial_number(value: str) -> str:
    return _normalize_digits(value)


def _format_sei_from_groups(groups: Sequence[str]) -> str:
    a, b, c, d, e = groups
    return f"{a}-{b}.{c}.{d}.{e}"


def _format_sei_candidate(raw: str) -> str:
    raw = raw.strip()
    if not raw:
        return ""
    match = PROCESSO_SEI_PATTERN.search(raw)
    if match:
        return match.group(0)
    tokens = [token for token in raw.split("_") if token]
    if len(tokens) >= 5 and all(token.isdigit() for token in tokens[:5]):
        candidate = f"{tokens[0]}-{tokens[1]}.{tokens[2]}.{tokens[3]}.{tokens[4]}"
        if PROCESSO_SEI_PATTERN.fullmatch(candidate):
            return candidate
    if "_" in raw:
        first, rest = raw.split("_", 1)
        candidate = f"{first}-{rest.replace('_', '.')}"
        if PROCESSO_SEI_PATTERN.fullmatch(candidate):
            return candidate
    loose = PROCESSO_SEI_LOOSE_PATTERN.search(raw)
    if loose:
        return _format_sei_from_groups(loose.groups())
    return ""


def _expected_sei_numbers(zip_name: str) -> tuple[Set[str], dict[str, str]]:
    stem = Path(zip_name).stem
    pieces = stem.split("_SEI_")
    candidates = []
    if len(pieces) == 2:
        candidates.extend(pieces)
    else:
        candidates.append(stem)
    normalized: Set[str] = set()
    display: dict[str, str] = {}
    for raw in candidates:
        formatted = _format_sei_candidate(raw)
        if formatted:
            norm = _normalize_sei_number(formatted)
            normalized.add(norm)
            display.setdefault(norm, formatted)
    return normalized, display


def _build_documents(sources: list[dict[str, str]]) -> list[DocumentText]:
    documents: list[DocumentText] = []
    for src in sources:
        text = src["text"]
        sei_numbers = _extract_sei_numbers(text)
        judicial_numbers = {_normalize_judicial_number(match) for match in PROCESSO_NUM_PATTERN.findall(text)}
        admin_numbers = _extract_admin_candidates(text)
        cnj_counts, cnj_context, cnj_display = _extract_cnj_metadata(text)
        importance = _document_importance(src["name"], text)
        documents.append(
            DocumentText(
                name=src["name"],
                text=text,
                sei_numbers=sei_numbers,
                judicial_numbers=judicial_numbers,
                admin_numbers=admin_numbers,
                cnj_counts=cnj_counts,
                cnj_context=cnj_context,
                cnj_display=cnj_display,
                importance=importance,
                bucket=src.get("bucket", DocumentBucket.OUTRO),
            )
        )
    return documents


def _extract_sei_numbers(text: str) -> Set[str]:
    numbers: Set[str] = set()
    for match in PROCESSO_SEI_PATTERN.finditer(text):
        if _has_cnj_tail(text, match.end()):
            continue
        numbers.add(_normalize_sei_number(match.group(0)))
    for loose in PROCESSO_SEI_LOOSE_PATTERN.finditer(text):
        formatted = _format_sei_from_groups(loose.groups())
        numbers.add(_normalize_sei_number(formatted))
    return numbers


def _has_cnj_tail(text: str, end_index: int) -> bool:
    tail = text[end_index : end_index + 5]
    return bool(re.match(r"\.\d{4}", tail))


def _extract_admin_candidates(text: str) -> Set[str]:
    numbers: Set[str] = set()
    for match in PROCESSO_ADMIN_PATTERN.findall(text):
        numbers.add(_normalize_digits(match))
    for match in PROCESSO_ADMIN_LEGACY_PATTERN.findall(text):
        digits = _normalize_digits(match)
        if len(digits) >= 7:
            numbers.add(digits)
    return numbers


def _numbers_match(a: str, b: str) -> bool:
    if not a or not b:
        return False
    shorter, longer = (a, b) if len(a) <= len(b) else (b, a)
    if len(shorter) < 6:
        return shorter == longer
    return longer.startswith(shorter)


def _matches_any(values: Set[str], reference: Set[str]) -> bool:
    for value in values:
        for ref in reference:
            if value == ref or _numbers_match(value, ref):
                return True
    return False


def _document_is_relevant(doc: DocumentText, context: ProcessContext) -> bool:
    if context.expected_sei and not context.accepted_docs:
        if doc.sei_numbers and _matches_any(doc.sei_numbers, context.expected_sei):
            return True
        return False

    if doc.sei_numbers:
        if _matches_any(doc.sei_numbers, context.expected_sei):
            return True
        if _matches_any(doc.sei_numbers, context.sei_numbers):
            return True
        return False

    has_anchor = context.has_anchor()

    if doc.admin_numbers:
        if context.admin_numbers & doc.admin_numbers:
            return True
        if not has_anchor:
            return True

    if doc.judicial_numbers:
        if _matches_any(doc.judicial_numbers, context.judicial_numbers):
            return True
        if not has_anchor:
            return True

    if not doc.has_ids:
        return has_anchor

    return False


def _missing_required_fields(bucket: DocumentBucket, result: ExtractionResult) -> bool:
    required = BUCKET_REQUIREMENTS.get(bucket)
    if not required:
        return False
    for field in required:
        if not result.data.get(field):
            return True
    return False


def _should_expand_bucket(bucket: DocumentBucket, context: ProcessContext, result: ExtractionResult) -> bool:
    if bucket == DocumentBucket.OUTRO:
        return False
    if not context.accepted_docs:
        return True
    return _missing_required_fields(bucket, result)


def _summarize_documents(context: ProcessContext, result: ExtractionResult) -> list[dict[str, object]]:
    summaries: list[dict[str, object]] = []
    if not context.accepted_docs:
        return summaries
    field_sources: dict[str, list[str]] = {}
    for field, source in result.sources.items():
        field_sources.setdefault(source, []).append(field)
    for doc in context.accepted_docs:
        fields = []
        for field in sorted(field_sources.get(doc.name, [])):
            meta = result.meta.get(field, {})
            fields.append(
                {
                    "field": field,
                    "value": result.data.get(field, ""),
                    "pattern": meta.get("pattern", ""),
                    "snippet": meta.get("snippet", ""),
                    "page": meta.get("page"),
                    "start": meta.get("start"),
                    "end": meta.get("end"),
                }
            )
        summaries.append({"name": doc.name, "bucket": doc.bucket.value, "fields": fields})
    return summaries


def _result_to_audit_entry(zip_name: str, result: ExtractionResult, run_id: str) -> dict[str, object]:
    documents = result.meta.get("_documents", [])
    bucket_lookup = {doc.get("name"): doc.get("bucket") for doc in documents}
    fields = []
    for field in sorted(result.data.keys()):
        value = result.data.get(field)
        if not value:
            continue
        meta = result.meta.get(field, {})
        source = result.sources.get(field, "")
        fields.append(
            {
                "field": field,
                "value": value,
                "source": source,
                "bucket": bucket_lookup.get(source, ""),
                "pattern": meta.get("pattern", ""),
                "snippet": meta.get("snippet", ""),
                "page": meta.get("page"),
                "start": meta.get("start"),
                "end": meta.get("end"),
            }
        )
    bucket_counts = result.meta.get("_bucket_usage", {}).get("counts", {})
    return {
        "run_id": run_id,
        "zip": zip_name,
        "zip_path": result.meta.get("_zip_path", ""),
        "bucket_counts": bucket_counts,
        "documents": documents,
        "fields": fields,
        "observations": list(result.observations),
    }


def _append_audit_entries(path: Path, items: list[tuple[str, ExtractionResult]], run_id: str) -> None:
    if not items:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as handle:
        for zip_name, result in items:
            entry = _result_to_audit_entry(zip_name, result, run_id)
            handle.write(json.dumps(entry, ensure_ascii=False))
            handle.write("\n")


def _extract_cnj_metadata(text: str) -> tuple[dict[str, int], dict[str, int], dict[str, str]]:
    counts: dict[str, int] = {}
    context_scores: dict[str, int] = {}
    display: dict[str, str] = {}
    lower_text = text.lower()
    for match in PROCESSO_NUM_PATTERN.finditer(text):
        raw = match.group(0)
        norm = _normalize_judicial_number(raw)
        if not norm:
            continue
        display.setdefault(norm, raw)
        counts[norm] = counts.get(norm, 0) + 1
        start = max(0, match.start() - 120)
        end = min(len(text), match.end() + 120)
        snippet = lower_text[start:end]
        bonus = 0
        for keyword in CNJ_CONTEXT_KEYWORDS:
            if keyword in snippet:
                bonus += 1
        context_scores[norm] = context_scores.get(norm, 0) + bonus
    return counts, context_scores, display


def _document_importance(name: str, text: str) -> int:
    priority, _ = document_priority(name, text)
    return max(1, 20 - priority)


def extract_from_text(text: str, combined: str, source_doc: str) -> ExtractionResult:
    res = ExtractionResult(data={})
    primary = text or combined
    lookup_text = primary if primary else combined
    if not lookup_text:
        res.observations.append("Sem texto legível no ZIP")
        return res

    lines = _prepare_lines(lookup_text)
    doc_origin = _classify_arbitration_doc(source_doc, lookup_text)

    processo_cnj = _sanitize_cnj(_find_first(PROCESSO_NUM_PATTERN, lookup_text))
    _set_field(res, "PROCESSO Nº", processo_cnj, source_doc, pattern="processo_regex", context_text=lookup_text)
    _set_field(res, "PROCESSO ADMIN. Nº", _extract_admin_number(lookup_text), source_doc, pattern="admin_regex", context_text=lookup_text)
    _set_field(res, "JUÍZO", _find_first(JUÍZO_PATTERN, lookup_text), source_doc, pattern="juizo_regex", context_text=lookup_text)
    _set_field(res, "COMARCA", _find_first(COMARCA_PATTERN, lookup_text), source_doc, pattern="comarca_regex", context_text=lookup_text)

    if not res.data.get("JUÍZO"):
        req_line = _line_value(lines, ("juízo", "vara"))
        if req_line:
            _set_field(res, "JUÍZO", req_line, source_doc, pattern="juizo_line", context_text=lookup_text, weight=0.9)
    if not res.data.get("JUÍZO"):
        juizo_requerente = _juizo_from_requerente(lines)
        if juizo_requerente:
            _set_field(res, "JUÍZO", juizo_requerente, source_doc, pattern="juizo_requerente", context_text=lookup_text, weight=0.85)

    if not res.data.get("COMARCA"):
        comarca = _extract_comarca(res.data.get("JUÍZO", ""))
        if not comarca:
            comarca = _extract_comarca(lookup_text)
        if comarca:
            _set_field(res, "COMARCA", comarca, source_doc, pattern="comarca_from_juizo", context_text=lookup_text, weight=0.9)

    promovente, promovido = _extract_partes(lines, lookup_text)
    if promovente:
        _set_field(res, "PROMOVENTE", promovente, source_doc, pattern="partes_regex", context_text=lookup_text)
    else:
        _set_field(res, "PROMOVENTE", _line_value(lines, PROMOVENTE_LABELS), source_doc, pattern="promovente_labels", context_text=lookup_text, weight=0.9)
    if promovido:
        _set_field(res, "PROMOVIDO", promovido, source_doc, pattern="partes_regex", context_text=lookup_text)
    else:
        _set_field(res, "PROMOVIDO", _line_value(lines, PROMOVIDO_LABELS), source_doc, pattern="promovido_labels", context_text=lookup_text, weight=0.9)

    perito_info = _extract_perito_info(lines)
    if perito_info.nome:
        _set_field(res, "PERITO", perito_info.nome, source_doc, pattern="perito_info", context_text=lookup_text)
    if perito_info.documento:
        _set_field(res, "CPF/CNPJ", perito_info.documento, source_doc, pattern="perito_info", context_text=lookup_text)
    if perito_info.especialidade:
        _set_field(res, "ESPECIALIDADE", perito_info.especialidade, source_doc, pattern="perito_info", context_text=lookup_text)
    else:
        _set_field(res, "ESPECIALIDADE", _line_value(lines, ESPECIALIDADE_LABELS), source_doc, pattern="especialidade_labels", context_text=lookup_text, weight=0.9)
    especie = _extract_especie_from_text(lines, lookup_text)
    if especie:
        _apply_species_mapping(res, especie, source_doc, context_text=lookup_text, weight=1.0)
    elif perito_info.especialidade or perito_info.profissao:
        alias_entry = _guess_species_from_specialty(perito_info)
        if alias_entry:
            _apply_species_mapping(
                res,
                alias_entry.get("DESCRICAO", ""),
                source_doc,
                context_text=lookup_text,
                weight=0.7,
                matched_entry=alias_entry,
            )

    fator = _line_value(lines, FATOR_LABELS)
    if not fator:
        fator = _find_after_labels(lookup_text, FATOR_LABELS, max_len=50)
    if fator and not res.data.get("Fator"):
        _set_field(res, "Fator", fator, source_doc, pattern="fator_label", context_text=lookup_text, weight=0.8)

    val_tab = _line_value(lines, VALOR_TABELA_LABELS)
    if not val_tab:
        val_tab = _find_after_labels(lookup_text, VALOR_TABELA_LABELS, max_len=80)
    if val_tab and not res.data.get("Valor Tabelado Anexo I - Tabela I"):
        _set_field(res, "Valor Tabelado Anexo I - Tabela I", val_tab, source_doc, pattern="valor_tabelado", context_text=lookup_text, weight=0.8)

    valor_arbitrado = _line_value(lines, VALOR_ARBITRADO_LABELS)
    if not valor_arbitrado:
        valor_arbitrado = _first_currency(lines, keywords=("honor", "perícia", "perito"))
    _record_arbitration_value(res, valor_arbitrado, source_doc, lookup_text, doc_origin=doc_origin)

    _set_field(res, "CHECAGEM", _line_value(lines, CHECAGEM_LABELS) or _find_after_labels(lookup_text, CHECAGEM_LABELS, 40), source_doc, pattern="checagem", context_text=lookup_text, weight=0.8)
    if doc_origin:
        weight_adiant = 1.1 if doc_origin == "CM" else 1.0
        data_adiant = _line_value(lines, DATA_ADIANTAMENTO_LABELS) or _find_after_labels(
            lookup_text, DATA_ADIANTAMENTO_LABELS, 40
        )
        _set_field(
            res,
            "DATA ADIANTAMENTO",
            data_adiant,
            source_doc,
            pattern=f"data_adiantamento_{doc_origin.lower()}",
            context_text=lookup_text,
            weight=weight_adiant,
        )
        checagem_adiant = _line_value(lines, CHECAGEM_ADIANT_LABELS) or _find_after_labels(
            lookup_text, CHECAGEM_ADIANT_LABELS, 40
        )
        _set_field(
            res,
            "CHECAGEM ADIANTAMENTO",
            checagem_adiant,
            source_doc,
            pattern=f"checagem_adiant_{doc_origin.lower()}",
            context_text=lookup_text,
            weight=weight_adiant,
        )
        _set_field(
            res,
            "Data da Autorização da Despesa",
            _line_value(lines, DATA_AUTORIZACAO_LABELS) or _find_after_labels(
                lookup_text, DATA_AUTORIZACAO_LABELS, 60
            ),
            source_doc,
            pattern=f"autorizacao_despesa_{doc_origin.lower()}",
            context_text=lookup_text,
            weight=weight_adiant,
        )
    else:
        _set_field(
            res,
            "Data da Autorização da Despesa",
            _line_value(lines, DATA_AUTORIZACAO_LABELS) or _find_after_labels(
                lookup_text, DATA_AUTORIZACAO_LABELS, 60
            ),
            source_doc,
            pattern="autorizacao_despesa",
            context_text=lookup_text,
            weight=0.7,
        )
    _set_field(res, "SALDO A RECEBER", _line_value(lines, SALDO_LABELS) or _find_after_labels(lookup_text, SALDO_LABELS, 60), source_doc, pattern="saldo", context_text=lookup_text, weight=0.7)

    requisicao = _find_label_date(lookup_text, ["data da requisição", "data do requerimento", "campina grande", "joão pessoa", "patos", "sousa"])
    _set_field(res, "DATA DA REQUISIÇÃO", requisicao, source_doc, pattern="data_requisicao", context_text=lookup_text)

    _set_field(res, "R$", res.data.get("VALOR ARBITRADO", ""), source_doc, pattern="valor_arbitrado", context_text=lookup_text)
    percent_value = _extract_percentage(lookup_text) if doc_origin else ""
    _set_field(
        res,
        "%",
        percent_value,
        source_doc,
        pattern=f"percentual_{doc_origin.lower()}" if doc_origin else "percentual",
        context_text=lookup_text,
        weight=0.6 if not doc_origin else (1.0 if doc_origin == "DE" else 1.1),
    )

    # Observações para campos críticos ausentes
    for critical in ("PROCESSO Nº", "PERITO", "VALOR ARBITRADO"):
        if not res.data.get(critical):
            res.observations.append(f"Sem {critical}")

    return res


def _find_first(pattern: re.Pattern, text: str) -> str:
    if not text:
        return ""
    match = pattern.search(text)
    return match.group(0) if match else ""


def _sanitize_cnj(value: str) -> str:
    if not value or not PROCESSO_NUM_PATTERN.fullmatch(value):
        return ""
    digits = re.sub(r"\D", "", value)
    if len(digits) != 20:
        return ""
    seq = digits[:7]
    dv = digits[7:9]
    ano = digits[9:13]
    segmento = digits[13]
    tribunal = digits[14:16]
    origem = digits[16:]
    base = int(f"{seq}{ano}{segmento}{tribunal}{origem}")
    remainder = base % 97
    calculated = 98 - remainder
    if calculated == 0:
        calculated = 1
    dv_calc = f"{calculated:02d}"
    return value if dv_calc == dv else ""


def _set_field(
    result: ExtractionResult,
    field: str,
    value: str,
    source_doc: str,
    pattern: str = "",
    snippet: str | None = None,
    page: int | None = None,
    start: int | None = None,
    end: int | None = None,
    context_text: str | None = None,
    weight: float = 1.0,
) -> None:
    if not value:
        return
    if (start is None or end is None) and context_text:
        start, end = _locate_value(context_text, value)
    candidate = {
        "value": value,
        "source": source_doc,
        "pattern": pattern,
        "snippet": snippet or _compute_snippet(context_text, start, end) or value,
        "page": page,
        "start": start,
        "end": end,
        "weight": weight,
    }
    result.candidates.setdefault(field, []).append(candidate)
    current_meta = result.meta.get(field)
    current_weight = current_meta.get("weight", 0) if current_meta else 0
    if not result.data.get(field) or weight >= current_weight:
        result.data[field] = value
        result.sources[field] = source_doc
        result.meta[field] = candidate


def _extract_admin_number(text: str) -> str:
    if not text:
        return ""
    match = PROCESSO_ADMIN_PATTERN.search(text)
    if match:
        return match.group(0)
    match = PROCESSO_ADMIN_LEGACY_PATTERN.search(text)
    if match:
        digits = _normalize_digits(match.group(0))
        return _format_admin_number(digits)
    return ""


def _format_admin_number(digits: str) -> str:
    if not digits:
        return ""
    if len(digits) <= 4:
        return digits
    if len(digits) == 10:
        return digits
    return f"{digits[:4]}.{digits[4:]}"


def _find_after_labels(raw_text: str, labels: Iterable[str], max_len: int = 200) -> str:
    """Captura texto logo após um rótulo (ex.: 'Espécie de Perícia: ...')."""
    if not raw_text:
        return ""
    lowered = raw_text.lower()
    for label in labels:
        lbl = label.lower()
        idx = lowered.find(lbl)
        if idx == -1:
            continue
        tail = raw_text[idx + len(lbl) :]
        if not tail:
            continue
        lines = tail.splitlines()
        if not lines:
            continue
        part = lines[0]
        # remove pontuação inicial
        part = re.sub(r"^[\\s:\\-–—]+", "", part).strip()
        if part:
            return part[:max_len]
    return ""


def _compute_snippet(text: str | None, start: int | None, end: int | None, radius: int = 80) -> str:
    if text is None or start is None or end is None:
        return ""
    begin = max(0, start - radius)
    finish = min(len(text), end + radius)
    return text[begin:finish].replace("\n", " ").strip()


def _locate_value(context_text: str | None, value: str) -> tuple[int | None, int | None]:
    if not context_text or not value:
        return None, None
    lower_text = context_text.lower()
    lower_value = value.lower()
    idx = lower_text.find(lower_value)
    if idx != -1:
        return idx, idx + len(lower_value)
    return None, None


@dataclass
class PeritoInfo:
    nome: str = ""
    documento: str = ""
    especialidade: str = ""
    profissao: str = ""


def _prepare_lines(text: str) -> list[str]:
    return [line.strip() for line in text.splitlines() if line.strip()]


def _line_value(lines: Sequence[str], labels: Iterable[str]) -> str:
    for line in lines:
        lower = line.lower()
        for label in labels:
            lbl = label.lower()
            if lower.startswith(lbl + ":"):
                return _clean_after_colon(line)
            if lbl in lower and ":" in line:
                before, after = line.split(":", 1)
                if lbl in before.lower():
                    return _clean_after_colon(line)
    return ""


def _clean_after_colon(line: str) -> str:
    after = line.split(":", 1)[1].strip()
    after = after.split(" – ")[0]
    after = after.split(" - ")[0]
    return after.strip()


def _extract_perito_info(lines: Sequence[str]) -> PeritoInfo:
    info = PeritoInfo()
    doc_text = "\n".join(lines)

    def candidate_lines(predicate):
        return [line for line in lines if predicate(line.lower())]

    preferred = candidate_lines(lambda l: "interessad" in l)
    if not preferred:
        preferred = candidate_lines(lambda l: "interessad" in l and "perit" in l)
    if not preferred:
        preferred = candidate_lines(lambda l: l.startswith("perito") or l.startswith("perita"))
    if not preferred:
        preferred = candidate_lines(lambda l: "perito" in l)

    for line in preferred:
        lower = line.lower()
        if "cpf" not in lower and ":" not in line:
            continue
        if "interessad" in lower and ":" in line:
            info.nome = _clean_after_colon(line)
            tail = line.split("–", 1)[1] if "–" in line else line.split("-", 1)[-1]
            info.especialidade = tail.split("-", 1)[0].strip()
            if "profiss" in lower:
                info.profissao = info.especialidade
        elif "cpf" in lower and "," in line:
            match = PERITO_PARAGRAPH_PATTERN.search(line)
            if match:
                info.nome = match.group("nome").strip()
                info.documento = match.group("cpf")
                info.especialidade = match.group("esp").strip()
        else:
            info.nome = line.split(" – ")[0].split(" - ")[0].strip()

        doc = CPF_PATTERN.search(line) or CNPJ_PATTERN.search(line)
        if doc:
            info.documento = doc.group(0)

        if not info.especialidade and "–" in line:
            info.especialidade = line.split("–", 1)[1].split("-", 1)[0].strip()
        elif not info.especialidade and "-" in line:
            info.especialidade = line.split("-", 1)[1].strip()
        if info.nome:
            break

    if not info.nome:
        match = PERITO_PARAGRAPH_PATTERN.search(doc_text)
        if match:
            info.nome = match.group("nome").strip()
            info.documento = match.group("cpf")
            info.especialidade = match.group("esp").strip()

    if not info.documento:
        doc = CPF_PATTERN.search(doc_text) or CNPJ_PATTERN.search(doc_text)
        if doc:
            info.documento = doc.group(0)
    if not info.profissao:
        prof_match = re.search(r"profiss[aã]o\s*[:\-]?\s*([A-Za-zÀ-ÿ ]+)", doc_text, re.IGNORECASE)
        if prof_match:
            info.profissao = prof_match.group(1).strip()
    return info


def _extract_partes(lines: Sequence[str], text: str) -> tuple[str, str]:
    match = PARTES_REGEX.search(text)
    if match:
        prom = _strip_juizo_tokens(_clean_entity(match.group("promovente")))
        prov = _strip_juizo_tokens(_clean_entity(match.group("promovido")))
        return prom, prov

    prom = _first_entity(lines, [PROMOVENTE_LABELS + ("autor", "parte autora", "exequente"), ("requerente",)])
    prov = _first_entity(lines, [PROMOVIDO_LABELS + ("réu", "executado", "parte ré"), ("requerido", "parte ré" )])
    if not prov:
        prov = _extract_promovido_from_phrases(text)
    return prom, prov


def _first_entity(lines: Sequence[str], label_groups: Iterable[Iterable[str]]) -> str:
    for labels in label_groups:
        value = _strip_juizo_tokens(_line_value(lines, labels))
        if value:
            return value
    return ""


def _juizo_from_requerente(lines: Sequence[str]) -> str:
    for line in lines:
        lower = line.lower()
        if "requerent" in lower and ("juízo" in lower or "vara" in lower):
            if ":" in line:
                return _clean_after_colon(line)
            return line.strip()
    return ""


def _extract_comarca(text: str) -> str:
    if not text:
        return ""
    match = re.search(r"(Comarca\s+(?:de|da|do|dos|das)\s+[A-Za-zÀ-ÿ ]+)", text, re.IGNORECASE)
    if match:
        value = match.group(1).strip()
        return value
    return ""


def _clean_entity(value: str) -> str:
    if not value:
        return ""
    value = value.split("CPF", 1)[0]
    value = value.split("CNPJ", 1)[0]
    value = value.replace("-", " ").strip()
    return value


def _strip_juizo_tokens(value: str) -> str:
    if not value:
        return ""
    pattern = re.compile(r"(Ju[ií]zo|Vara|Comarca).*$", re.IGNORECASE)
    match = pattern.search(value)
    if match:
        value = value[:match.start()].strip(" -–:;/")
    return value


def _extract_promovido_from_phrases(text: str) -> str:
    if not text:
        return ""
    for pattern in (PROMOVIDO_LABEL_PATTERN, PROMOVIDO_EM_FACE_PATTERN, PROMOVIDO_CONTRA_PATTERN):
        match = pattern.search(text)
        if match:
            candidate = match.group(1).strip()
            candidate = candidate.split(",")[0].split(";")[0].split("\n")[0]
            candidate = _clean_entity(candidate)
            candidate = _strip_juizo_tokens(candidate)
            if candidate:
                return candidate
    return ""


def _extract_especie_from_text(lines: Sequence[str], text: str) -> str:
    specie = _line_value(lines, ESPECIE_LABELS)
    if specie:
        return specie
    specie = _line_value(lines, NATUREZA_LABELS)
    if specie:
        return specie
    specie = _find_after_labels(text, ESPECIE_LABELS)
    if specie:
        return specie
    specie = _find_after_labels(text, NATUREZA_LABELS)
    if specie:
        return specie
    match = ESPECIE_INLINE_PATTERN.search(text)
    if match:
        return match.group(1).strip()
    match = NATUREZA_HONORARIOS_PATTERN.search(text)
    if match:
        return match.group(1).strip()
    match = NATUREZA_SERVICO_PATTERN.search(text)
    if match:
        return match.group(1).strip()
    return ""


def _guess_species_from_specialty(info: PeritoInfo) -> dict[str, str] | None:
    for text in (info.especialidade, info.profissao):
        entry = _match_alias(text)
        if entry:
            return entry
    return None


def _match_honorarios_entry(label: str) -> dict[str, str] | None:
    if not label or not HONORARIOS_INDEX:
        return None
    norm = _normalize_key(label)
    if not norm:
        return None
    if norm in HONORARIOS_INDEX:
        return HONORARIOS_INDEX[norm]
    candidates = difflib.get_close_matches(norm, HONORARIOS_INDEX.keys(), n=1, cutoff=0.75)
    if candidates:
        return HONORARIOS_INDEX[candidates[0]]
    for key, row in HONORARIOS_INDEX.items():
        if norm in key or key in norm:
            return row
    return None


def _match_alias(text: str | None) -> dict[str, str] | None:
    if not text or not HONORARIOS_ALIAS:
        return None
    norm = _normalize_key(text)
    if not norm:
        return None
    for alias in HONORARIOS_ALIAS:
        if any(keyword in norm for keyword in alias["keywords"]):
            return alias["entry"]
    return None


def _format_currency_value(value: str | float) -> str:
    if isinstance(value, str):
        try:
            value = float(value.replace(",", "."))
        except Exception:
            return ""
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _classify_arbitration_doc(source_doc: str, context_text: str | None) -> str:
    """Retorna 'CM', 'DE' ou ''."""
    name = (source_doc or "").lower()
    text = (context_text or "").lower() if context_text else ""
    if "assessoria do conselho da magistratura" in text and "certid" in text:
        return "CM"
    if "conselho da magistratura" in text or "conselho da magistratura" in name:
        return "CM"
    if "certidão" in name and ("magistratura" in name or "cm" in name):
        return "CM"
    if "cm" in name and "certida" in name:
        return "CM"
    if "diretoria especial" in text or "diesp" in text:
        return "DE"
    if "despacho" in name:
        return "DE"
    return ""


def _record_arbitration_value(
    result: ExtractionResult,
    value: str,
    source_doc: str,
    context_text: str | None,
    doc_origin: str | None = None,
) -> None:
    doc_type = doc_origin or _classify_arbitration_doc(source_doc, context_text)
    if doc_type == "CM":
        _set_field(
            result,
            "VALOR ARBITRADO - CM",
            value,
            source_doc,
            pattern="valor_arbitrado_cm",
            context_text=context_text,
            weight=1.2,
        )
        _set_field(
            result,
            "VALOR ARBITRADO",
            value,
            source_doc,
            pattern="valor_arbitrado_cm",
            context_text=context_text,
            weight=1.2,
        )
    elif doc_type == "DE":
        _set_field(
            result,
            "VALOR ARBITRADO - DE",
            value,
            source_doc,
            pattern="valor_arbitrado_de",
            context_text=context_text,
            weight=1.1,
        )
        _set_field(
            result,
            "VALOR ARBITRADO",
            value,
            source_doc,
            pattern="valor_arbitrado_de",
            context_text=context_text,
            weight=1.1,
        )
    else:
        _set_field(
            result,
            "VALOR ARBITRADO",
            value,
            source_doc,
            pattern="valor_arbitrado",
            context_text=context_text,
            weight=1.0,
        )


def _apply_species_mapping(
    result: ExtractionResult,
    especie: str,
    source_doc: str,
    context_text: str | None = None,
    weight: float = 1.0,
    matched_entry: dict[str, str] | None = None,
) -> None:
    if not especie:
        return
    especie_clean = especie.strip(" \t-–:;")
    if not especie_clean:
        return
    _set_field(
        result,
        "ESPÉCIE DE PERÍCIA",
        especie_clean,
        source_doc,
        pattern="especie_label",
        context_text=context_text,
        weight=weight,
    )
    entry = matched_entry or _match_honorarios_entry(especie_clean)
    if not entry:
        entry = _match_alias(especie_clean)
    if not entry:
        return
    especie_oficial = entry.get("DESCRICAO", especie_clean)
    tabela_source = source_doc or "tabela_honorarios"
    if especie_oficial:
        _set_field(
            result,
            "ESPÉCIE DE PERÍCIA",
            especie_oficial,
            tabela_source,
            pattern="tabela_honorarios",
            weight=weight + 0.1,
        )
    fator = entry.get("ID", "")
    if fator and not result.data.get("Fator"):
        _set_field(
            result,
            "Fator",
            fator,
            tabela_source,
            pattern="tabela_honorarios",
            weight=weight + 0.05,
        )
    valor = entry.get("VALOR", "")
    if valor and not result.data.get("Valor Tabelado Anexo I - Tabela I"):
        valor_fmt = _format_currency_value(valor)
        if valor_fmt:
            _set_field(
                result,
                "Valor Tabelado Anexo I - Tabela I",
                valor_fmt,
                tabela_source,
                pattern="tabela_honorarios",
                weight=weight + 0.05,
            )


def _first_currency(lines: Sequence[str], keywords: Iterable[str] | None = None) -> str:
    pattern = re.compile(r"R\$\s*[\d\.]+,\d{2}")
    for line in lines:
        lower = line.lower()
        if "r$" not in lower:
            continue
        if keywords and not any(keyword in lower for keyword in keywords):
            continue
        match = pattern.search(line)
        if match:
            return match.group(0)
    return ""


def _find_label_date(text: str, labels: Iterable[str]) -> str:
    lines = _prepare_lines(text)
    candidate = _line_value(lines, labels)
    if candidate:
        parsed = _parse_date(candidate)
        return parsed
    match = DATE_PATTERN.search(text)
    if match:
        parsed = _parse_date(match.group(0))
        return parsed or match.group(0)
    return ""


def _parse_date(raw: str) -> str:
    try:
        dt = date_parser.parse(raw, dayfirst=True, fuzzy=True)
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return ""


def _extract_percentage(text: str) -> str:
    patterns = [
        r"(?i)percentual\s*[:\-]?\s*(\d{1,3}(?:,\d{1,2})?)\s*%",
        r"(?i)percentual\s*[:\-]?\s*(\d{1,3}(?:,\d{1,2})?)",
        r"\b(\d{1,3}(?:,\d{1,2})?)\s*%",
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if not m:
            continue
        value = m.group(1).replace(",", ".")
        try:
            num = float(value)
        except ValueError:
            continue
        if 0 <= num <= 100:
            return f"{num:g}%"
    return ""


def _is_valid_cpf(digits: str) -> bool:
    digits = re.sub(r"\D", "", digits)
    if len(digits) != 11 or len(set(digits)) == 1:
        return False

    def dv(s: str, start: int) -> str:
        soma = sum(int(d) * w for d, w in zip(s, range(start, 1, -1)))
        r = 11 - (soma % 11)
        return "0" if r >= 10 else str(r)

    return dv(digits[:9], 10) == digits[9] and dv(digits[:10], 11) == digits[10]


def _is_valid_cnpj(digits: str) -> bool:
    digits = re.sub(r"\D", "", digits)
    if len(digits) != 14 or len(set(digits)) == 1:
        return False

    def calc(pos: int) -> str:
        soma = 0
        pesos = list(range(pos, 1, -1)) + list(range(9, 1, -1))
        for d, w in zip(digits[: pos - 1], pesos):
            soma += int(d) * w
        r = soma % 11
        return "0" if r < 2 else str(11 - r)

    return calc(13) == digits[12] and calc(14) == digits[13]


def _is_valid_doc(doc: str) -> bool:
    digits = re.sub(r"\D", "", doc or "")
    if len(digits) == 11:
        return _is_valid_cpf(digits)
    if len(digits) == 14:
        return _is_valid_cnpj(digits)
    return False


def _validate_date(value: str) -> str:
    if not value:
        return ""
    try:
        dt = date_parser.parse(value, dayfirst=True, fuzzy=True)
    except Exception:
        return ""
    if dt.year < 2000 or dt.year > 2026:
        return ""
    return dt.strftime("%d/%m/%Y")


def _validate_currency(value: str, min_v: float = 10.0, max_v: float = 500_000.0) -> str:
    if not value:
        return ""
    m = re.search(r"R\$\s*([0-9\.\s]{1,15},\d{2})", value)
    if not m:
        return ""
    num = float(m.group(1).replace(".", "").replace(" ", "").replace(",", "."))
    if num < min_v or num > max_v:
        return ""
    return f"R$ {num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _validate_percentage(value: str) -> str:
    if not value:
        return ""
    m = re.match(r"^\s*([0-9]{1,3}(?:[.,][0-9]{1,2})?)\s*%?\s*$", value)
    if not m:
        return ""
    num = float(m.group(1).replace(",", "."))
    if 0 <= num <= 100:
        return f"{num:g}%"
    return ""


def _validate_result(result: ExtractionResult, context: ProcessContext, zip_name: str, full_text: str) -> None:
    """Pós-processamento para aumentar especificidade; descarta suspeitos e anota observações."""
    obs_add = result.observations.append

    def _norm_sei(v: str) -> str:
        return _normalize_digits(v or "")

    # Processo deve bater com o ZIP (se houver referência)
    proc = result.data.get("PROCESSO Nº", "")
    if proc and context.expected_sei and _norm_sei(proc) not in context.expected_sei:
        obs_add(f"Processo divergente do ZIP: {proc}")

    # CPF/CNPJ válido
    doc = result.data.get("CPF/CNPJ", "")
    if doc and not _is_valid_doc(doc):
        obs_add(f"Documento suspeito: {doc}")
        result.data["CPF/CNPJ"] = ""

    # Datas em faixa
    for key in ("DATA DA REQUISIÇÃO", "DATA ADIANTAMENTO", "Data da Autorização da Despesa"):
        v = result.data.get(key, "")
        fixed = _validate_date(v)
        if v and not fixed:
            obs_add(f"Data suspeita em {key}: {v}")
        result.data[key] = fixed

    # Valores monetários plausíveis
    for key in ("VALOR ARBITRADO", "R$", "SALDO A RECEBER"):
        v = result.data.get(key, "")
        fixed = _validate_currency(v)
        if v and not fixed:
            obs_add(f"Valor suspeito em {key}: {v}")
        result.data[key] = fixed

    # Percentual
    pct_existing = result.data.get("%", "")
    pct_valid = _validate_percentage(pct_existing)
    if pct_existing and not pct_valid:
        obs_add(f"Percentual suspeito: {pct_existing}")
    if not pct_valid:
        pct_valid = _extract_percentage(full_text or "")
    result.data["%"] = pct_valid

    # Promovido ausente
    if not result.data.get("PROMOVIDO"):
        obs_add("Promovido ausente; revisar manualmente")


def process_zip(zip_path: Path) -> ExtractionResult:
    sources, combined = gather_texts(zip_path)
    documents = _build_documents(sources)
    expected_sei, expected_display = _expected_sei_numbers(zip_path.name)
    context = ProcessContext(expected_sei=expected_sei, expected_sei_display=expected_display)
    result = ExtractionResult()
    if not documents and not combined:
        result.observations.append("Nenhum documento legível no ZIP")
        return result

    accepted_texts: list[str] = []
    bucket_counts: dict[str, int] = {bucket.value: 0 for bucket in BUCKET_ORDER}
    documents_by_bucket: dict[DocumentBucket, list[DocumentText]] = {bucket: [] for bucket in BUCKET_ORDER}
    for doc in documents:
        documents_by_bucket.setdefault(doc.bucket, []).append(doc)

    for index, bucket in enumerate(BUCKET_ORDER):
        docs = documents_by_bucket.get(bucket, [])
        for doc in docs:
            if _document_is_relevant(doc, context):
                context.register(doc)
                accepted_texts.append(doc.text)
                partial = extract_from_text(doc.text, doc.text, doc.name)
                result.update_from(partial, doc.name)
                bucket_counts[bucket.value] = bucket_counts.get(bucket.value, 0) + 1
            else:
                context.skipped_docs.append(doc.name)

        is_last_bucket = index == len(BUCKET_ORDER) - 1
        if is_last_bucket:
            break
        if not _should_expand_bucket(bucket, context, result):
            break

    if not context.accepted_docs:
        if context.expected_sei:
            result.observations.append("Sem documento compatível com o processo SEI no ZIP")
        elif not documents:
            result.observations.append("Nenhum documento legível no ZIP")

    fallback_text = "\n".join(accepted_texts)
    if not result.data.get("PROCESSO Nº"):
        if fallback_text:
            fallback = extract_from_text(fallback_text, fallback_text, "combined")
            result.update_from(fallback, "combined")
        elif not context.accepted_docs and not context.expected_sei and combined:
            fallback = extract_from_text(combined, combined, "combined")
            result.update_from(fallback, "combined")

    if context.skipped_docs and context.accepted_docs:
        skipped = ", ".join(context.skipped_docs[:3])
        extra = len(context.skipped_docs) - 3
        msg = f"Documentos ignorados por divergência: {skipped}"
        if extra > 0:
            msg += f" (+{extra})"
        result.observations.append(msg)

    if context.accepted_docs:
        _select_primary_cnj(result, context)
    _apply_admin_fallback(result, context, zip_path.name)
    _fill_requisition_date(result, context)
    _fill_species_from_laudos(result, context.accepted_docs)
    _ensure_honorarios_completion(result)
    text_for_validation = "\n".join(accepted_texts) or combined
    _validate_result(result, context, zip_path.name, text_for_validation)

    if not sources:
        result.observations.append("Nenhum documento legível no ZIP")
    result.meta.setdefault("_bucket_usage", {"counts": bucket_counts})
    result.meta["_zip_path"] = str(zip_path)
    result.meta["_documents"] = _summarize_documents(context, result)
    return result


def _select_primary_cnj(result: ExtractionResult, context: ProcessContext) -> None:
    totals: dict[str, dict[str, object]] = {}
    for doc in context.accepted_docs:
        importance = max(1, doc.importance)
        for norm, count in doc.cnj_counts.items():
            entry = totals.setdefault(
                norm,
                {
                    "display": doc.cnj_display.get(norm, norm),
                    "weighted": 0,
                    "context": 0,
                    "raw": 0,
                    "sources": set(),
                },
            )
            entry["weighted"] += count * importance
            entry["context"] += doc.cnj_context.get(norm, 0) * importance
            entry["raw"] += count
            entry["sources"].add(doc.name)

    if not totals:
        return

    def _score(item: tuple[str, dict[str, object]]) -> tuple[int, int]:
        data = item[1]
        return (data["weighted"] * 2 + data["context"], data["raw"])

    primary_norm, primary_data = max(totals.items(), key=_score)
    primary_display = primary_data["display"]
    if primary_display:
        sources = primary_data["sources"]
        source_doc = next(iter(sources)) if sources else "doc"
        _set_field(
            result,
            "PROCESSO Nº",
            primary_display,
            source_doc,
            pattern="primary_cnj",
            weight=1.1,
        )

    additional = [data["display"] for norm, data in totals.items() if norm != primary_norm and data["display"]]
    if additional:
        msg = "CNJs adicionais mencionados: " + ", ".join(sorted(set(additional)))
        if msg not in result.observations:
            result.observations.append(msg)


def _apply_admin_fallback(result: ExtractionResult, context: ProcessContext, zip_name: str) -> None:
    current = result.data.get("PROCESSO ADMIN. Nº", "").strip()
    display = _primary_sei_display(context)
    if display:
        if current and _normalize_digits(current) != _normalize_digits(display):
            obs = f"Processo adm. citado nos documentos: {current}"
            if obs not in result.observations:
                result.observations.append(obs)
        _set_field(
            result,
            "PROCESSO ADMIN. Nº",
            display,
            f"zip:{zip_name}",
            pattern="admin_context",
            weight=0.9,
        )
    elif current:
        formatted = _format_admin_number(_normalize_digits(current))
        _set_field(
            result,
            "PROCESSO ADMIN. Nº",
            formatted,
            f"zip:{zip_name}",
            pattern="admin_formatted",
            weight=0.8,
        )
    else:
        fallback = Path(zip_name).stem
        formatted = _format_sei_candidate(fallback)
        if formatted:
            _set_field(
                result,
                "PROCESSO ADMIN. Nº",
                formatted,
                f"zip:{zip_name}",
                pattern="zip_fallback",
                weight=0.6,
            )


def _primary_sei_display(context: ProcessContext) -> str:
    if context.expected_sei_display:
        return next(iter(context.expected_sei_display.values()))
    if context.expected_sei:
        norm = next(iter(context.expected_sei))
        return norm
    return ""


def _fill_requisition_date(result: ExtractionResult, context: ProcessContext) -> None:
    if result.data.get("DATA DA REQUISIÇÃO"):
        return
    date, doc = _extract_requisition_date_from_docs(context.accepted_docs)
    if date:
        source = doc.name if doc else "context"
        context_text = doc.text if doc else None
        _set_field(
            result,
            "DATA DA REQUISIÇÃO",
            date,
            source,
            pattern="data_requisicao_doc",
            context_text=context_text,
            weight=0.85,
        )


def _fill_species_from_laudos(result: ExtractionResult, documents: List[DocumentText]) -> None:
    if result.data.get("ESPÉCIE DE PERÍCIA") and result.data.get("Fator") and result.data.get("Valor Tabelado Anexo I - Tabela I"):
        return
    for doc in documents:
        if "laudo" not in doc.name.lower():
            continue
        specie = _extract_especie_from_text(_prepare_lines(doc.text), doc.text)
        if specie:
            _apply_species_mapping(result, specie, doc.name, context_text=doc.text, weight=0.85)
            if result.data.get("Fator") and result.data.get("Valor Tabelado Anexo I - Tabela I"):
                break


def _ensure_honorarios_completion(result: ExtractionResult) -> None:
    specie = result.data.get("ESPÉCIE DE PERÍCIA")
    if not specie:
        return
    source = result.sources.get("ESPÉCIE DE PERÍCIA", "tabela_honorarios")
    entry = _match_honorarios_entry(specie) or _match_alias(specie)
    _apply_species_mapping(result, specie, source, context_text=None, weight=0.7, matched_entry=entry)


def _extract_requisition_date_from_docs(docs: List[DocumentText]) -> tuple[str, DocumentText | None]:
    for doc in docs[:3]:
        date = _date_near_cities(doc.text)
        if date:
            return date, doc
        date = _date_near_requisition(doc.text)
        if date:
            return date, doc
    return "", None


def _date_near_cities(text: str) -> str:
    lower = text.lower()
    for city in CITY_KEYWORDS:
        start = 0
        while True:
            idx = lower.find(city, start)
            if idx == -1:
                break
            snippet = text[idx : idx + 160]
            date = _first_date_in_text(snippet)
            if date:
                return date
            start = idx + len(city)
    return ""
    start = 0
    while True:
        idx = lower.find(" joao pessoa", start)
        if idx == -1:
            break
        snippet = text[idx : idx + 160]
        date = _first_date_in_text(snippet)
        if date:
            return date
        start = idx + 11
    return ""


def _first_date_in_text(text: str) -> str:
    for match in DATE_NUMERIC_PATTERN.finditer(text):
        parsed = _parse_date(match.group(0))
        if parsed:
            return parsed
    for match in DATE_PATTERN.finditer(text):
        parsed = _parse_date(match.group(0))
        if parsed:
            return parsed
    return ""


def _date_near_requisition(text: str) -> str:
    lower = text.lower()
    for match in DATE_NUMERIC_PATTERN.finditer(text):
        snippet = lower[max(0, match.start() - 80) : match.end() + 80]
        if _snippet_has_keyword(snippet):
            parsed = _parse_date(match.group(0))
            if parsed:
                return parsed
    for match in DATE_PATTERN.finditer(text):
        snippet = lower[max(0, match.start() - 80) : match.end() + 80]
        if _snippet_has_keyword(snippet):
            parsed = _parse_date(match.group(0))
            if parsed:
                return parsed
    return ""


def _snippet_has_keyword(snippet: str) -> bool:
    return any(keyword in snippet for keyword in REQUISITION_KEYWORDS)


def _norm_name(value: str | None) -> str:
    if not value:
        return ""
    value = value.strip().lower()
    return "".join(c for c in unicodedata.normalize("NFD", value) if unicodedata.category(c) != "Mn")


def _load_perito_catalog() -> tuple[set[str], dict[str, str]]:
    """Retorna (nomes_normalizados, mapa_nome_normalizado->CPF) do catálogo externo."""
    global _PERITO_NAME_SET
    names: set[str] = set()
    name_to_cpf: dict[str, str] = {}
    try:
        if _PERITO_CATALOG_PATH.exists():
            import pandas as pd

            df = pd.read_csv(_PERITO_CATALOG_PATH)
            if "PERITO" in df.columns:
                for _, row in df.iterrows():
                    n = _norm_name(str(row.get("PERITO", "")))
                    if not n:
                        continue
                    names.add(n)
                    cpf = str(row.get("CPF/CNPJ", "")).strip()
                    if cpf:
                        name_to_cpf[n] = cpf
    except Exception:
        names = set()
        name_to_cpf = {}
    _PERITO_NAME_SET = names
    return names, name_to_cpf


def _ensure_comarca_from_juizo(result: "ExtractionResult") -> None:
    juizo = (result.data.get("JUÍZO") or "").lower()
    comarca = result.data.get("COMARCA", "") or ""
    if not juizo:
        return
    if not comarca:
        # padrão "comarca de xxxx"
        import re

        m = re.search(r"comarca\s+de\s+([a-zçãáâêéíóõú\s]+)", juizo)
        if m:
            comarca_val = m.group(1).strip().title()
            result.data["COMARCA"] = comarca_val
            return
    # casos "... Vara ... da capital"
    if "capital" in juizo:
        result.data["COMARCA"] = "João Pessoa"


def _scrub_perito_conflicts(result: "ExtractionResult") -> None:
    """Remove perito em campos de partes e ajusta CPF do perito pelo catálogo."""
    perito_names, name_to_cpf = _load_perito_catalog()

    # limpar partes que coincidam com perito
    for field in ("PROMOVENTE", "PROMOVIDO"):
        value = result.data.get(field, "")
        if value and _norm_name(value) in perito_names:
            result.data[field] = ""
            result.observations.append(f"{field} coincidia com nome de perito; valor removido")

    # ajustar CPF do perito usando catálogo
    perito_nome = result.data.get("PERITO", "") or ""
    if perito_nome:
        norm = _norm_name(perito_nome)
        cat_cpf = name_to_cpf.get(norm, "")
        if cat_cpf:
            current_cpf = result.data.get("CPF/CNPJ", "") or ""
            if current_cpf != cat_cpf:
                result.data["CPF/CNPJ"] = cat_cpf
                result.observations.append("CPF do perito ajustado pelo catálogo externo")

    # garantir comarca a partir do juízo
    _ensure_comarca_from_juizo(result)

def _norm_name(value: str | None) -> str:
    if not value:
        return ""
    value = value.strip().lower()
    return "".join(c for c in unicodedata.normalize("NFD", value) if unicodedata.category(c) != "Mn")


def _load_perito_catalog() -> set[str]:
    global _PERITO_NAME_SET
    if _PERITO_NAME_SET is not None:
        return _PERITO_NAME_SET
    names: set[str] = set()
    try:
        if _PERITO_CATALOG_PATH.exists():
            import pandas as pd

            df = pd.read_csv(_PERITO_CATALOG_PATH)
            if "PERITO" in df.columns:
                names = {_norm_name(n) for n in df["PERITO"].dropna()}
    except Exception:
        names = set()
    _PERITO_NAME_SET = names
    return names


def _scrub_perito_conflicts(result: "ExtractionResult") -> None:
    """Resolve conflitos e enriquece dados de perito usando catálogo externo."""
    perito_df = None
    try:
        perito_df = pd.read_csv(_PERITO_CATALOG_PATH) if _PERITO_CATALOG_PATH.exists() else None
    except Exception:
        perito_df = None
    perito_names = _load_perito_catalog()

    # 1) Promovente/Promovido não podem ser perito
    for field in ("PROMOVENTE", "PROMOVIDO"):
        value = result.data.get(field, "")
        if not value:
            continue
        if _norm_name(value) in perito_names:
            result.data[field] = ""
            result.observations.append(f"{field} coincidia com nome de perito; valor removido")

    # 2) Enriquecer CPF/perito a partir do catálogo quando faltante ou divergente
    perito_nome = result.data.get("PERITO", "") or ""
    perito_cpf = result.data.get("CPF/CNPJ", "") or ""
    if perito_df is not None and perito_nome:
        norm = _norm_name(perito_nome)
        matches = perito_df[perito_df["PERITO"].apply(_norm_name) == norm]
        if not matches.empty:
            cat_cpf = str(matches.iloc[0].get("CPF/CNPJ", "")).strip()
            if cat_cpf and cat_cpf != perito_cpf:
                result.data["CPF/CNPJ"] = cat_cpf
                result.observations.append("CPF do perito ajustado pelo catálogo externo")


def _load_existing_parquet_names(parquet_dir: Path) -> set[str]:
    """Retorna nomes de ZIP que já possuem parquet salvo."""
    if not parquet_dir.exists():
        return set()
    return {p.stem for p in parquet_dir.glob("*.parquet")}


def process_and_save_parquet(zip_name: str, resolved_path: str, parquet_dir: str) -> str:
    """Worker: processa um arquivo e salva parquet (1 arquivo por ZIP)."""
    path = Path(resolved_path)
    result = process_zip(path)
    _scrub_perito_conflicts(result)
    pdir = Path(parquet_dir)
    pdir.mkdir(parents=True, exist_ok=True)
    tmp_path = pdir / f"{zip_name}.parquet.tmp"
    final_path = pdir / f"{zip_name}.parquet"
    df = pd.DataFrame([result.to_row(0, zip_name)], columns=COLUMNS)
    df.to_parquet(tmp_path, index=False)
    tmp_path.replace(final_path)
    return zip_name, result


def consolidate_parquets(parquet_dir: Path, excel_path: Path) -> list[Path]:
    """Consolida todos os parquets existentes em um Excel único.

    Retorna a lista de parquets corrompidos/ignorados.
    """
    files = sorted(parquet_dir.glob("*.parquet"))
    if not files:
        return []
    dfs = []
    bad_files = []
    for f in files:
        try:
            dfs.append(pd.read_parquet(f))
        except Exception:
            bad_files.append(f)
    if bad_files:
        _log(f"Aviso: {len(bad_files)} parquet(s) corrompido(s) ignorado(s): {[p.name for p in bad_files]}")
    if not dfs:
        return bad_files
    df_all = pd.concat(dfs, ignore_index=True)

    # Fallback para VALOR ARBITRADO: CM > DE (somente valor monetário)
    money_re = re.compile(r"r\$\s*[0-9]{1,3}(?:\.[0-9]{3})*,?\d{2}", re.IGNORECASE)

    def _money(val: str | None) -> str:
        if not isinstance(val, str):
            return ""
        m = money_re.search(val)
        return m.group(0).strip() if m else ""

    if "VALOR ARBITRADO" not in df_all.columns:
        df_all["VALOR ARBITRADO"] = ""
    mask_empty = df_all["VALOR ARBITRADO"].isna() | (df_all["VALOR ARBITRADO"] == "")
    if "VALOR ARBITRADO - CM" in df_all.columns:
        df_all.loc[mask_empty, "VALOR ARBITRADO"] = df_all.loc[mask_empty, "VALOR ARBITRADO - CM"].apply(_money)
    mask_empty = df_all["VALOR ARBITRADO"].isna() | (df_all["VALOR ARBITRADO"] == "")
    if "VALOR ARBITRADO - DE" in df_all.columns:
        df_all.loc[mask_empty, "VALOR ARBITRADO"] = df_all.loc[mask_empty, "VALOR ARBITRADO - DE"].apply(_money)
    # Garante colunas e renumera
    for c in COLUMNS:
        if c not in df_all.columns:
            df_all[c] = ""
    df_all = df_all[COLUMNS]
    df_all["Nº DE PERÍCIAS"] = range(1, len(df_all) + 1)
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    df_all.to_excel(excel_path, index=False)
    return bad_files


def _append_results_to_workbook(
    wb: Workbook,
    start_index: int,
    results: list[tuple[str, ExtractionResult]],
) -> None:
    pericias = wb["Pericias"]
    pend = wb["Pendencias"] if "Pendencias" in wb.sheetnames else wb.create_sheet("Pendencias")
    if pend.max_row == 0:
        pend.append(COLUMNS)
    fontes_headers = [
        "Nº DE PERÍCIAS",
        "Campo",
        "Valor",
        "Documento Fonte",
        "Pattern/Heurística",
        "Snippet",
        "Page",
        "Start",
        "End",
        "ZIP",
    ]
    if "Fontes" in wb.sheetnames:
        fontes = wb["Fontes"]
        if fontes.max_row == 0:
            fontes.append(fontes_headers)
    else:
        fontes = wb.create_sheet("Fontes")
        fontes.append(fontes_headers)

    cand_headers = [
        "Nº DE PERÍCIAS",
        "Campo",
        "Valor",
        "Peso",
        "Fonte",
        "Pattern",
        "Snippet",
        "Start",
        "End",
        "ZIP",
    ]
    if "Candidatos" in wb.sheetnames:
        candidatos = wb["Candidatos"]
        if candidatos.max_row == 0:
            candidatos.append(cand_headers)
    else:
        candidatos = wb.create_sheet("Candidatos")
        candidatos.append(cand_headers)

    obs_idx = COLUMNS.index("OBSERVACOES")

    for offset, (zip_name, result) in enumerate(results, start=1):
        row_index = start_index + offset
        row = result.to_row(row_index, zip_name)
        pericias.append([_safe_excel_value(v) for v in row])
        if row[obs_idx]:
            pend.append([_safe_excel_value(v) for v in row])
        for field, source in sorted(result.sources.items()):
            meta = result.meta.get(field, {})
            fontes.append(
                [
                    _safe_excel_value(f"{row_index:02d}"),
                    _safe_excel_value(field),
                    _safe_excel_value(result.data.get(field, "")),
                    _safe_excel_value(meta.get("source", source)),
                    _safe_excel_value(meta.get("pattern", "")),
                    _safe_excel_value(meta.get("snippet", "")),
                    _safe_excel_value(meta.get("page", "")),
                    _safe_excel_value(meta.get("start", "")),
                    _safe_excel_value(meta.get("end", "")),
                    _safe_excel_value(zip_name),
                ]
            )
        for field, entries in result.candidates.items():
            for entry in entries:
                candidatos.append(
                    [
                        _safe_excel_value(f"{row_index:02d}"),
                        _safe_excel_value(field),
                        _safe_excel_value(entry.get("value", "")),
                        _safe_excel_value(entry.get("weight", "")),
                        _safe_excel_value(entry.get("source", "")),
                        _safe_excel_value(entry.get("pattern", "")),
                        _safe_excel_value(entry.get("snippet", "")),
                        _safe_excel_value(entry.get("start", "")),
                        _safe_excel_value(entry.get("end", "")),
                        _safe_excel_value(zip_name),
                    ]
                )


def write_excel(results: list[tuple[str, ExtractionResult]], output: Path, append: bool = False) -> None:
    if append and output.exists():
        try:
            wb = load_workbook(output)
            if "Pericias" not in wb.sheetnames:
                append = False
            else:
                start_index = max(0, wb["Pericias"].max_row - 1)
                _append_results_to_workbook(wb, start_index, results)
                wb.save(output)
                return
        except Exception:
            append = False

    wb = Workbook()
    ws = wb.active
    ws.title = "Pericias"
    ws.append(COLUMNS)
    pend = wb.create_sheet("Pendencias")
    pend.append(COLUMNS)
    fontes = wb.create_sheet("Fontes")
    fontes.append([
        "Nº DE PERÍCIAS",
        "Campo",
        "Valor",
        "Documento Fonte",
        "Pattern/Heurística",
        "Snippet",
        "Page",
        "Start",
        "End",
        "ZIP",
    ])
    candidatos = wb.create_sheet("Candidatos")
    candidatos.append([
        "Nº DE PERÍCIAS",
        "Campo",
        "Valor",
        "Peso",
        "Fonte",
        "Pattern/Heurística",
        "Snippet",
        "Start",
        "End",
        "ZIP",
    ])
    _append_results_to_workbook(wb, 0, results)
    wb.save(output)


def append_single_result(
    output: Path,
    zip_name: str,
    result: ExtractionResult,
) -> None:
    # Persistência incremental segura: Parquet por ZIP
    parquet_dir = output.parent / "parquet"
    parquet_dir.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame([result.to_row(0, zip_name)], columns=COLUMNS)
    df.to_parquet(parquet_dir / f"{zip_name}.parquet", index=False)


def main() -> None:
    parser = argparse.ArgumentParser(description="Extrai dados dos despachos do SEI a partir de arquivos locais.")
    parser.add_argument("--zip-dir", type=Path, help="Diretório com os arquivos ZIP.")
    parser.add_argument("--pdf-dir", type=Path, action="append", help="Diretório contendo PDFs avulsos (pode repetir).")
    parser.add_argument("--txt-dir", type=Path, action="append", help="Diretório contendo arquivos .txt (pode repetir).")
    parser.add_argument("--output", type=Path, default=Path("relatorio-pericias.xlsx"), help="Caminho do arquivo XLSX de saída.")
    parser.add_argument("--limit", type=int, default=None, help="Limita a quantidade de arquivos processados.")
    parser.add_argument(
        "--skip-existing",
        action="store_true",
        help="Se o arquivo de saída já existir, ignora os registros já presentes e acrescenta somente os novos.",
    )
    parser.add_argument("--workers", type=int, default=24, help="Número de processos em paralelo (default=24).")
    parser.add_argument("--run-id", help="Identificador personalizado da execução.")
    parser.add_argument("--resume", help="Retoma a execução indicada (run-id).")
    parser.add_argument(
        "--checkpoint-interval",
        type=int,
        default=25,
        help="Quantidade de arquivos processados antes de salvar a planilha/checkpoint.",
    )
    parser.add_argument(
        "--log-retention-days",
        type=int,
        default=30,
        help="Remove logs/checkpoints com mais de N dias (0 desativa).",
    )
    args = parser.parse_args()

    if args.resume and args.run_id:
        parser.error("Use apenas --run-id ou --resume, não ambos.")

    if args.resume:
        run_id = args.resume
        state = _load_state(run_id)
    else:
        run_id = args.run_id or _generate_run_id()
        if args.log_retention_days:
            _cleanup_old_logs(args.log_retention_days)
        state = {"run_id": run_id, "created_at": datetime.now().isoformat()}

    log_path = _setup_logger(run_id)
    audit_path = LOG_DIR / f"{run_id}.sources.jsonl"
    _log(f"Executando extração (run-id={run_id}) - log: {log_path}")
    zip_paths: list[Path] = []
    pdf_paths: list[Path] = []
    txt_paths: list[Path] = []

    if args.zip_dir:
        zip_dir = args.zip_dir.expanduser()
        if not zip_dir.exists():
            raise SystemExit(f"Diretório não encontrado: {zip_dir}")
        zip_paths.extend(sorted(path for path in zip_dir.glob("*.zip")))
    for directory in args.pdf_dir or []:
        dir_path = directory.expanduser()
        if not dir_path.exists():
            raise SystemExit(f"Diretório não encontrado: {dir_path}")
        pdf_paths.extend(sorted(path for path in dir_path.glob("*.pdf")))
    for directory in args.txt_dir or []:
        dir_path = directory.expanduser()
        if not dir_path.exists():
            raise SystemExit(f"Diretório não encontrado: {dir_path}")
        txt_paths.extend(sorted(path for path in dir_path.glob("*.txt")))

    total_size = sum(p.stat().st_size for p in zip_paths + pdf_paths + txt_paths)
    _log(f"Pré-flight ok | ZIPs: {len(zip_paths)} PDFs: {len(pdf_paths)} TXTs: {len(txt_paths)} | Tamanho total: {total_size/1e6:.1f} MB")

    total_expected = len(zip_paths) + len(pdf_paths) + len(txt_paths)

    loose_paths = pdf_paths + txt_paths

    if not zip_paths and not loose_paths:
        raise SystemExit("Informe ao menos um diretório via --zip-dir, --pdf-dir ou --txt-dir.")

    prepared_inputs, temp_dir = resolve_input_paths(zip_paths=zip_paths, pdf_paths=loose_paths, limit=None)
    if not prepared_inputs:
        _log("Nenhum arquivo suportado encontrado.")
        return

    output = args.output.expanduser()
    output.parent.mkdir(parents=True, exist_ok=True)
    state_processed = set(state.get("processed_files", []))
    state["output"] = str(output)

    parquet_dir = output.parent / "parquet"
    processed_zips: set[str] = set()
    if args.skip_existing and parquet_dir.exists():
        processed_zips = _load_existing_parquet_names(parquet_dir)
        if processed_zips:
            _log(f"{len(processed_zips)} registro(s) já presentes (parquet); serão ignorados.")

    processed_set = set(processed_zips)
    processed_set.update(state_processed)

    remaining_inputs = [entry for entry in prepared_inputs if entry.original.name not in processed_set]
    skipped = len(prepared_inputs) - len(remaining_inputs)
    if skipped:
        _log(f"Pulando {skipped} arquivo(s) já presentes no relatório.")

    if args.limit:
        remaining_inputs = remaining_inputs[: args.limit]

    if not remaining_inputs:
        _log("Nenhum arquivo pendente. Nada a fazer.")
        return

    checkpoint_interval = max(1, args.checkpoint_interval)
    total_to_process = len(remaining_inputs)
    _log(f"Lista de trabalho: {total_to_process} arquivo(s) | workers={args.workers} | checkpoint a cada {checkpoint_interval}.")

    file_sizes = {prepared.original.name: Path(prepared.resolved).stat().st_size for prepared in remaining_inputs}
    checkpoint_bytes = 0
    checkpoint_start = time.time()
    _print_header(run_id, log_path, total_to_process)

    def consolidate_checkpoint(final: bool = False) -> None:
        nonlocal checkpoint_bytes, checkpoint_start
        state["processed_files"] = sorted(state_processed)
        state["last_update"] = datetime.now().isoformat()
        state["completed"] = final
        _save_state(run_id, state)
        # consolida todos parquets existentes em Excel
        bad = consolidate_parquets(parquet_dir, output)
        if bad:
            bad_files_total.update([p.name for p in bad])
        elapsed = time.time() - checkpoint_start
        mbps = (checkpoint_bytes / 1e6) / elapsed if elapsed > 0 else 0
        _log(f"Checkpoint salvo ({len(state_processed)} registros) | {checkpoint_bytes/1e6:.1f} MB em {elapsed:.1f}s ({mbps:.2f} MB/s).")
        checkpoint_bytes = 0
        checkpoint_start = time.time()

    bad_files_total: set[str] = set()

    try:
        t0 = time.time()
        with ProcessPoolExecutor(max_workers=max(1, args.workers)) as executor:
            futures = {
                executor.submit(
                    process_and_save_parquet,
                    prepared.original.name,
                    str(prepared.resolved),
                    str(parquet_dir),
                ): prepared.original.name
                for prepared in remaining_inputs
            }
            completed = 0
            for future in as_completed(futures):
                name, result = future.result()
                completed += 1
                _log_progress(completed, total_to_process, name)
                processed_set.add(name)
                state_processed.add(name)
                checkpoint_bytes += file_sizes.get(name, 0)
                if completed % checkpoint_interval == 0:
                    consolidate_checkpoint(final=False)
        elapsed = time.time() - t0
        mb = total_size / 1e6 if total_size else 0
        if elapsed > 0 and mb:
            _log(f"Processamento paralelo concluído: {completed}/{total_to_process} arquivos | {mb/elapsed:.2f} MB/s")
    finally:
        if temp_dir:
            temp_dir.cleanup()

    consolidate_checkpoint(final=True)
    _finish_progress()
    _log(f"Relatório salvo/atualizado em {output}")
    if bad_files_total:
        _log(f"Aviso final: {len(bad_files_total)} parquet(s) corrompido(s) foram ignorados: {sorted(bad_files_total)}")


if __name__ == "__main__":
    main()
